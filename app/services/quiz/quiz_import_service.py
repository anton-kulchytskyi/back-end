"""Service for importing quizzes from Excel files.

Expected Excel format (row 1 = header):
  quiz_title | quiz_description | question | answer_1 | is_correct_1 |
  answer_2 | is_correct_2 | answer_3 | is_correct_3 | answer_4 | is_correct_4

Rules:
  - Multiple rows sharing the same quiz_title form one quiz with multiple questions.
  - answer_3/is_correct_3 and answer_4/is_correct_4 columns are optional.
  - If a quiz with the same title already exists in the company it is updated
    (all questions replaced). Otherwise a new quiz is created.
  - is_correct values are accepted as Excel boolean (TRUE/FALSE), string
    "true"/"false", or integer 1/0.
"""

import io

from openpyxl import load_workbook
from pydantic import ValidationError

from app.core.exceptions import ServiceException
from app.core.logger import logger
from app.core.unit_of_work import AbstractUnitOfWork
from app.models import User
from app.schemas import (
    QuizAnswerCreateRequest,
    QuizCreateRequest,
    QuizImportResponse,
    QuizImportResult,
    QuizQuestionCreateRequest,
    QuizUpdateRequest,
)
from app.services.companies.permission_service import PermissionService
from app.services.quiz.quiz_service import QuizService


class QuizImportService:
    """Import quizzes from an Excel (.xlsx) file."""

    def __init__(
        self,
        uow: AbstractUnitOfWork,
        permission_service: PermissionService,
        quiz_service: QuizService,
    ):
        self._uow = uow
        self._permission_service = permission_service
        self._quiz_service = quiz_service

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    async def import_quizzes(
        self,
        company_id: int,
        current_user: User,
        file_bytes: bytes,
    ) -> QuizImportResponse:
        """Parse the Excel file and create or update quizzes in the company."""

        await self._permission_service.require_admin(company_id, current_user.id)

        quizzes = self._parse_excel(file_bytes)

        if not quizzes:
            raise ServiceException("No valid quiz data found in the file")

        results: list[QuizImportResult] = []

        for quiz_data in quizzes:
            async with self._uow:
                existing = await self._uow.quiz.get_by_title_and_company(
                    title=quiz_data.title,
                    company_id=company_id,
                )

            if existing:
                await self._quiz_service.update_quiz(
                    quiz_id=existing.id,
                    current_user=current_user,
                    data=QuizUpdateRequest(
                        title=quiz_data.title,
                        description=quiz_data.description,
                        questions=quiz_data.questions,
                    ),
                )
                results.append(
                    QuizImportResult(
                        title=quiz_data.title,
                        action="updated",
                        quiz_id=existing.id,
                    )
                )
                logger.info(
                    f"Quiz '{quiz_data.title}' updated via import in company {company_id}"
                )
            else:
                created = await self._quiz_service.create_quiz(
                    company_id=company_id,
                    current_user=current_user,
                    data=quiz_data,
                )
                results.append(
                    QuizImportResult(
                        title=quiz_data.title,
                        action="created",
                        quiz_id=created.id,
                    )
                )
                logger.info(
                    f"Quiz '{quiz_data.title}' created via import in company {company_id}"
                )

        return QuizImportResponse(imported=len(results), results=results)

    # ------------------------------------------------------------------
    # Excel parsing
    # ------------------------------------------------------------------

    def _parse_excel(self, file_bytes: bytes) -> list[QuizCreateRequest]:
        """Parse raw Excel bytes into a list of QuizCreateRequest objects."""

        wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
        ws = wb.active

        # Preserve insertion order: title → {description, questions}
        quizzes: dict[str, dict] = {}

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue

            quiz_title = str(row[0]).strip()
            quiz_description = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            question_title = str(row[2]).strip() if len(row) > 2 and row[2] else ""

            if not quiz_title or not question_title:
                continue

            answers = self._parse_answers(row)
            if not answers:
                continue

            if quiz_title not in quizzes:
                quizzes[quiz_title] = {
                    "description": quiz_description,
                    "questions": [],
                }

            quizzes[quiz_title]["questions"].append(
                QuizQuestionCreateRequest(title=question_title, answers=answers)
            )

        result: list[QuizCreateRequest] = []
        for title, data in quizzes.items():
            try:
                result.append(
                    QuizCreateRequest(
                        title=title,
                        description=data["description"],
                        questions=data["questions"],
                    )
                )
            except ValidationError as exc:
                first_error = exc.errors()[0]["msg"]
                raise ServiceException(
                    f"Invalid data for quiz '{title}': {first_error}"
                ) from exc

        return result

    @staticmethod
    def _parse_answers(row: tuple) -> list[QuizAnswerCreateRequest]:
        """Extract up to 4 answers from a data row (columns 3-10)."""
        answers: list[QuizAnswerCreateRequest] = []

        for i in range(4):
            col_text = 3 + i * 2
            col_flag = 4 + i * 2

            if col_text >= len(row):
                break

            text = row[col_text]
            if not text:
                continue

            raw_flag = row[col_flag] if col_flag < len(row) else False

            if isinstance(raw_flag, bool):
                is_correct = raw_flag
            elif isinstance(raw_flag, str):
                is_correct = raw_flag.strip().upper() == "TRUE"
            elif isinstance(raw_flag, (int, float)):
                is_correct = bool(raw_flag)
            else:
                is_correct = False

            answers.append(
                QuizAnswerCreateRequest(text=str(text).strip(), is_correct=is_correct)
            )

        return answers
